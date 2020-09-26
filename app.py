"""
        #variables

print('* 8 8 *'*10)
price = 10
rating =4.9
shivam ='sachdeva'
print(price,rating)
is_published =True
print(is_published)
print(shivam)

name = input('what is your name')
color = input('what is your favourite color')
print(name+' likes '+color)

  #type conversion

birth_year =input('your date of birth: ')
age = 2020 - int(birth_year)
print(type(age))
print(age)

weight = input('what is your weight in pounds')
weightkg = float(weight)*.45
weightkg = str(weightkg)
print(type(weightkg))   #important

   #strings

course = "python's beginner course"
coursea = '''
this is something crazy to understand 
i am starting python from now 
hope i will enjoy this 
thanks for watching my videos 
i am very thankful to you 
'''
print(course[2])
print(course[-1])
print(course[0:6])
print(course[0:])
print(course[:6])
print(course[1:-1])
print(coursea)

        #formattedstrings

first = 'john'
last = 'smith'
msg = f'{first}[{last}] is a coder'
print(msg)

    #string methods
course = "python's beginner course"
print(len(course))
print(course.upper())
print(course)                                           # IMP
print(course.lower())
print(course.find('course'))
print(course.replace('p','j'))
print('python' in course)

    #arthmetic operations

x =10
x+=1
print(x)
print(10/3)
print(10//3)
print(10**3)
print(10%3)
print(3+2*2**2)

    #math functions

import math
print(round(2.9),
abs(-2.9),
math.ceil(2.9),
math.floor(2.9))

  # if statenent

credit=True
income=100
if credit:
    print(f'they need to put down by {(income*10)/100}')
else:
    print(f'they need to put down by {(income*20)}/100')
print('enjoy your day and have a nice day ahead')

       #logical operators

name =input('what is your name')
if len(name)<3:
    print('name must be atleast 3 characters')
elif len(name)>=50:
    print('Name cannot be 50 characters long')
else :
    print('Name Looks Good')

  #weight converter project
weight =input('weight :')
convert =input('(L)bs or (K)g:')
if convert.upper()=='L':                                                                 #convert=='L' or convert =='l':
    weight= float(weight)*.45
    print(f'you are {weight} kgs')

else:
    weight=float(weight)/.45
    print(f'you are {weight} pounds')

       #LOOPs
#while Loop

secret=8
guess_count=0
while guess_count<=2:                                    #IMP INDENTATION ERROR MAY COMR
  guess=input('Guess:')
  if guess=='8':
     print('you win!!')
     break
  guess_count+=1
else :                                                    #IMP
    print('sorry you failed!')

command=''
started=False
                                                                                    #i=0
while True:                                                                #IMP
    command = input('>').lower()


    if command=='start':

                                                                                  #i += 1
        if started :                                                                       #i>=2
            print("car is already running")
                                                                                    # i=0
        else :
            started=True
            print('ready to go')
    elif command=='stop':

                                                                                        #i-=1
        if  not started:                                                                           #i<0
            print('already stopped')
                                                                                    #i=0
        else :
            started=False
            print('stopped')
    elif command=='help':
        print('helping')
                                                                                     # i = 0
    elif command=='quit':
        break
    else :
        print('i dont understand')
                                                                                 #  i = 0
#for loops

prices=[10,20,30]
t=0
for sum in prices:
   t+=sum
print(t)

for item in range (2,12,2):
    print(item)

#nested loops

for x in range(3):
    for y in range(3):
        print(f'({x},{y})')


numbers = [5,2,5,2,2]
for x in numbers:
  output=''
  for y in range(x):
      output+='x'
  print(output)

numbers = [1,1,1,1,5]
for x in numbers:
  output=''
  for y in range(x):
      output+='x'
  print(output)

             #LISTS

num=[4,5,4,8,6,8]
max=num[0]
for x in num:
    if x>max:
        max=x
print(x)

    #2D LISTS
matrix=[
      [2,3,4],
      [5,6,7]
]
print(matrix[0][1])
print(matrix)
for row in matrix:
    for item in row:
         print(item)

  #LIST METHODS OR FUNCTIONS

x=[5,7,5,7,7,7,7,78,7,7,7,5,5,5,5,5,7,7,75]
a=[]
for y in x:
        if y not in a:
            a.append(y)
print(a)

  #DICTIONSRIES

x=input('phone:')
convert={
    "1":"one",
    "2":"two",
    "3":"three",
    "4":"four",
    "5":"five",
    "6":"six"
}
output=""
for y in x:
    output+=convert.get(y,"!")+" "
print(output)

    #emoji converter

output=""
message = input('>')
messages=message.split(' ')
emojis = {
    ":)":"😊",
    ":(":"😒"}
for y in messages:
    output+=emojis.get(y,y)+" "
print(output)

    #functions
emojis = {
    ":)":"😊",
    ":(":"😒"}

def ec(command):
    preoutput = ""
    commands=command.split(" ")
    for y in commands:
        preoutput+=emojis.get(y,y) + " "
    return preoutput


message= input('<')
print(ec(message))

 # tackle with errors

try:
    age= int(input('enter your age :'))
    print(age)
except ValueError:
    print('invalid age')


 #  classes,constructors

class Person:
    def __init__(self,name):
      self.name=name
    def talk(self):
        print(f'hi {self.name} who are you to talk to me like this')

john=Person("john smith")
print(john.name)
john.talk()

      #inheritance
        #modules
import converters
print(converters.k_to_l(7))                                   #calling without a prefix i.e converters.k_to_l

from utils import find_max
print(find_max([1,2,3,5,78]))
#max() functions is also present in python which returns the greatest value

      #packages
import ecommerce.shipping
ecommerce.shipping.calcshipping()

import random


class Dice:
    def roll(self):
        x=random.randint(1,6)
        y=random.randint(1,6)
        return (x,y)

p=Dice()
print(p.roll())

#files and directories
from pathlib import Path
path=Path("")
print(path.exists())
for file in path.glob('*'):
   print(file)

      #EXCEL SPREADSHEETS

import openpyxl as xl
from openpyxl.chart import BarChart,Reference

wb=xl.load_workbook('transactions.xlsx') #will load the xlsx file and return file as a object
sheet = wb['Sheet1']   #to access sheets in excel file in this case it has only one sheet
#cell = sheet['a1']    #or cell=sheeet.cell(1,1) .This is to access cells,in both cases we are passing the coordinates of cell,in first case coordinates are according to excel sheet
print(cell.value)   #to print the value of cell
print(sheet.max_row)   #give the no. of rows in sheet

for row in range(2,sheet.max_row+1):     #iterating over the rows
    cell = sheet.cell(row,3)
    corrected_price = cell.value*0.9
    corrected_price_cell=sheet.cell(row,4)     #will add a new cell to the respective row
    corrected_price_cell.value=corrected_price
values=Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)
chart=BarChart()
chart.add_data(values)        #for more informations read the documentations openpyxl
sheet.add_chart(chart,'e2')    #take what and where
wb.save('transactions2.xlsx')

#to organize our cod we can use functions to use this

"""
      # MACHINE LEARNING PROJECT





























